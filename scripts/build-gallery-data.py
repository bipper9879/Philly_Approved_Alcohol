import argparse
import json
import re
from pathlib import Path
from urllib.parse import quote

from openpyxl import load_workbook


IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp"}


def normalize_location_key(value: str) -> str:
    text = (value or "").strip().lower()
    text = text.replace("&", " and ")
    return re.sub(r"[^a-z0-9]", "", text)


def is_valid_site_code(value: str) -> bool:
    text = (value or "").strip()
    if not text or text.lower() == "site code":
        return False
    # Accept broad operational formats (PHI-xxxxxx, DC-123456, numeric, etc.)
    return bool(re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9\-_]{1,}", text))


def to_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize_job_token(value):
    text = to_text(value)
    if not text:
        return None

    number = parse_number(text)
    if number is not None:
        integer_value = int(number)
        if abs(number - integer_value) < 1e-9 and integer_value > 0:
            return str(integer_value)

    token = re.sub(r"\s+", "", text).upper()
    if re.fullmatch(r"[A-Z0-9][A-Z0-9\-_]{1,}", token):
        return token
    return None


def parse_job_numbers(col_d_label, col_d_value, col_e_label, col_e_value):
    jobs = []
    d_value = parse_number(col_d_value)
    e_value = parse_number(col_e_value)
    d_job = normalize_job_token(col_d_label)
    e_job = normalize_job_token(col_e_label)

    if d_job and d_value is not None and 1 <= d_value <= 10:
        jobs.append(d_job)
    if e_job and e_value is not None and 1 <= e_value <= 10:
        jobs.append(e_job)

    return sorted(set(jobs))


def extract_hyperlink(cell):
    link = getattr(cell, "hyperlink", None)
    if not link:
        return None
    target = getattr(link, "target", None)
    if target:
        return str(target).strip()
    location = getattr(link, "location", None)
    if location:
        return str(location).strip()
    value = getattr(cell, "value", None)
    if isinstance(value, str):
        match = re.match(r'(?i)=\s*HYPERLINK\("([^"]+)"\s*,', value)
        if match:
            return match.group(1).strip()
    return None


def normalize_street_view_url(raw_value, lat, lon):
    text = to_text(raw_value)
    if text and re.match(r"^https?://", text, flags=re.IGNORECASE):
        return text
    if lat is not None and lon is not None:
        return f"https://www.google.com/maps/@?api=1&map_action=pano&viewpoint={lat},{lon}"
    return text or None


def find_header_row(ws):
    max_scan = min(100, ws.max_row or 100)
    for row in range(1, max_scan + 1):
        site_code = to_text(ws.cell(row=row, column=1).value)
        location = to_text(ws.cell(row=row, column=3).value)
        if site_code.lower() == "site code" and location.lower() == "location":
            return row
    raise RuntimeError("Could not find header row containing 'Site Code' and 'Location'.")


def read_workbook_metadata(workbook_path: Path, worksheet_name: str):
    wb = load_workbook(workbook_path, data_only=False, read_only=True)
    try:
        ws = wb[worksheet_name] if worksheet_name else wb.worksheets[0]
        header_row = find_header_row(ws)
        filter_label = to_text(ws.cell(row=header_row, column=4).value) or "Column D"
        col_d_label = to_text(ws.cell(row=header_row, column=4).value) or "Column D"
        col_e_label = to_text(ws.cell(row=header_row, column=5).value) or "Column E"

        by_exact = {}
        by_normalized = {}

        for row in range(header_row + 1, (ws.max_row or header_row) + 1):
            site_code = to_text(ws.cell(row=row, column=1).value)
            if not is_valid_site_code(site_code):
                continue

            location = to_text(ws.cell(row=row, column=3).value)
            if not location:
                continue

            col_d_raw = ws.cell(row=row, column=4).value
            col_e_raw = ws.cell(row=row, column=5).value
            street_view_cell = ws.cell(row=row, column=6)
            lat = parse_number(ws.cell(row=row, column=7).value)
            lon = parse_number(ws.cell(row=row, column=8).value)
            street_view = normalize_street_view_url(
                extract_hyperlink(street_view_cell) or street_view_cell.value,
                lat,
                lon,
            )
            job_numbers = parse_job_numbers(col_d_label, col_d_raw, col_e_label, col_e_raw)
            reviewer_eligible = len(job_numbers) > 0
            payload = {
                "siteCode": site_code,
                "streetViewUrl": street_view,
                "lat": lat,
                "lon": lon,
                "jobNumbers": job_numbers,
                "jobColumnLabels": {"d": col_d_label, "e": col_e_label},
                "reviewerEligible": reviewer_eligible,
                "reviewerFilterLabel": filter_label,
                "reviewerFilterValue": parse_number(col_d_raw),
            }

            by_exact[location] = payload
            by_normalized[normalize_location_key(location)] = payload

        return by_exact, by_normalized
    finally:
        wb.close()


def has_images(folder: Path) -> bool:
    return any(
        child.is_file() and child.suffix.lower() in IMAGE_EXTENSIONS
        for child in folder.iterdir()
    )


def discover_location_folders(index_files_root: Path, max_depth: int = 6):
    discovered = []

    def walk(folder: Path, depth: int):
        if depth > max_depth:
            return
        if has_images(folder):
            discovered.append(folder)
            return

        children = sorted([p for p in folder.iterdir() if p.is_dir()], key=lambda p: p.name.lower())
        for child in children:
            walk(child, depth + 1)

    for root_child in sorted([p for p in index_files_root.iterdir() if p.is_dir()], key=lambda p: p.name.lower()):
        walk(root_child, 1)

    return discovered


def build_image_url(index_files_root: Path, file_path: Path, url_prefix: str) -> str:
    parts = [quote(part, safe="") for part in file_path.relative_to(index_files_root).parts]
    return f"{url_prefix}/{'/'.join(parts)}"


def build_locations(index_files_root: Path, by_exact, by_normalized, url_prefix: str):
    grouped = {}

    for folder in discover_location_folders(index_files_root):
        key = normalize_location_key(folder.name)
        if not key:
            continue

        bucket = grouped.setdefault(key, {"location": folder.name, "folders": []})
        bucket["folders"].append(folder)

    locations = []
    for bucket in sorted(grouped.values(), key=lambda item: item["location"].lower()):
        location_name = bucket["location"]
        images = []
        seen_urls = set()
        seen_names = set()

        for folder in bucket["folders"]:
            for file_path in sorted([f for f in folder.iterdir() if f.is_file()], key=lambda p: p.name.lower()):
                if file_path.suffix.lower() not in IMAGE_EXTENSIONS:
                    continue
                image_url = build_image_url(index_files_root, file_path, url_prefix)
                image_name_key = file_path.name.lower()
                if image_name_key in seen_names:
                    continue
                if image_url in seen_urls:
                    continue
                seen_names.add(image_name_key)
                seen_urls.add(image_url)
                images.append({"name": file_path.name, "url": image_url})

        cover_image_name = None
        if images:
            explicit_cover = next((img for img in images if img["name"].lower() == "cover.jpg"), None)
            cover_image_name = explicit_cover["name"] if explicit_cover else images[0]["name"]

        metadata = by_exact.get(location_name) or by_normalized.get(normalize_location_key(location_name), {})
        primary_folder = bucket["folders"][0]
        primary_folder_rel = "/".join(quote(part, safe="") for part in primary_folder.relative_to(index_files_root).parts)
        locations.append(
            {
                "siteCode": metadata.get("siteCode"),
                "streetViewUrl": metadata.get("streetViewUrl"),
                "lat": metadata.get("lat"),
                "lon": metadata.get("lon"),
                "jobNumbers": metadata.get("jobNumbers") or [],
                "jobColumnLabels": metadata.get("jobColumnLabels") or {"d": "Column D", "e": "Column E"},
                "reviewerEligible": bool(metadata.get("reviewerEligible")),
                "reviewerFilterLabel": metadata.get("reviewerFilterLabel"),
                "reviewerFilterValue": metadata.get("reviewerFilterValue"),
                "location": location_name,
                "folderName": primary_folder.name,
                "folderUrl": f"{url_prefix}/{primary_folder_rel}/",
                "coverImageName": cover_image_name,
                "images": images,
            }
        )

    return locations


def main():
    parser = argparse.ArgumentParser(description="Build gallery-data.json directly from workbook + image folders.")
    parser.add_argument("--workbook-path", required=True, help="Absolute path to workbook (.xlsx/.xlsm/.xls)")
    parser.add_argument("--worksheet-name", default="", help="Worksheet name (default: first worksheet)")
    parser.add_argument("--repo-root", default=str(Path(__file__).resolve().parent.parent), help="Repository root path")
    parser.add_argument("--images-root", default="", help="Image root folder containing location subfolders")
    parser.add_argument("--city", required=True, help="City tag written into each location")
    parser.add_argument("--email", default="bipper9879@hotmail.com", help="Email in gallery-data.json payload")
    parser.add_argument("--output-path", default="", help="Output gallery-data path (default: <repo-root>/gallery-data.json)")
    parser.add_argument("--skip-image-source-write", action="store_true", help="Do not update data/image-source.json")
    parser.add_argument("--url-prefix", default="images", help="URL prefix used for generated image URLs")
    args = parser.parse_args()

    repo_root = Path(args.repo_root).resolve()
    workbook_path = Path(args.workbook_path).resolve()
    index_files_root = Path(args.images_root).resolve() if args.images_root else (repo_root / "index_files")
    output_path = Path(args.output_path).resolve() if args.output_path else (repo_root / "gallery-data.json")
    image_source_config_path = repo_root / "data" / "image-source.json"

    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook file not found: {workbook_path}")
    if not index_files_root.exists():
        raise FileNotFoundError(f"Missing index_files folder: {index_files_root}")

    by_exact, by_normalized = read_workbook_metadata(workbook_path, args.worksheet_name)
    url_prefix = (args.url_prefix or "images").strip().strip("/")
    if not url_prefix:
        url_prefix = "images"
    locations = build_locations(index_files_root, by_exact, by_normalized, url_prefix)
    city = (args.city or "").strip()
    if not city:
        raise RuntimeError("City is required. Pass --city with a non-empty value.")
    for location in locations:
        location["city"] = city

    payload = {
        "email": args.email,
        "issueUrlBase": "https://github.com/bipper9879/buildPortfolio/issues/new",
        "locations": locations,
    }

    output_path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")

    if not args.skip_image_source_write:
        image_source_config_path.parent.mkdir(parents=True, exist_ok=True)
        image_source_config_path.write_text(
            json.dumps(
                {
                    "imagesRootPath": str(index_files_root),
                    "city": city,
                    "workbookPath": str(workbook_path),
                    "worksheetName": (args.worksheet_name or "").strip(),
                },
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )

    matched = sum(1 for loc in locations if loc.get("siteCode"))
    unmatched = len(locations) - matched
    print(f"gallery-data.json rebuilt. {len(locations)} locations written.")
    print(f"  Site codes matched: {matched}  |  Unmatched: {unmatched}")
    if unmatched:
        for loc in locations:
            if not loc.get("siteCode"):
                print(f"  No site code: {loc['location']}")


if __name__ == "__main__":
    main()
